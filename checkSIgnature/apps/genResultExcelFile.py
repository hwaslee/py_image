import openpyxl
import os.path
import logging


excelPath = "F:\\ws_py_windwos\\hkfmi\\venv\\files"
excelFilename = "check_signature5.xlsx"
headers = ['File명', '설계번호', '고객ID', '계약자 성명', '계약자 서명', '예금주 성명', '예금주 서명']

def initExcelFile():
    excelFullname = os.path.join(excelPath, excelFilename)

    if not os.path.exists(excelFullname):
        ## workbook = xlsxwriter.Workbook(excelFullname)
        ## worksheet = workbook.add_worksheet()
        workbook = openpyxl.Workbook()
        ## worksheet = workbook.create_sheet("Signature")
        worksheet = workbook.active

        worksheet.append(headers)
        workbook.save(excelFullname)
        workbook.close()

'''
def getMaxRow22():
    excelFullname = os.path.join(excelPath, excelFilename)

    ## workbook = xlsxwriter.Workbook(excelFullname)
    ## worksheet = workbook.add_worksheet()
    workbook = openpyxl.load_workbook(excelFullname)
    worksheet = workbook.worksheets[0]

    return worksheet.max_row
'''


def writeToExcel(desNo, custId, results, srcFilename):
    # data = ['abc.jpg', '222', '333', 'O', 'O', 'X', 'X']
    excelFullname = os.path.join(excelPath, excelFilename)

    workbook = openpyxl.load_workbook(excelFullname)
    worksheet = workbook.worksheets[0]

    data = []
    index = 0
    data.insert(index, srcFilename)
    index += 1
    data.insert(index, desNo)
    index += 1
    data.insert(index, custId)
    index += 1

    for result in results:
        data.insert(index, result)
        index += 1

    worksheet.append(data)
    workbook.save(excelFullname)
    workbook.close()
    logging.debug("----------------- Excel process completed ---------------------")


if __name__ == '__main__':
    initExcelFile()
    writeToExcel()
    writeToExcel()
